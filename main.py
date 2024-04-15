from ftplib import FTP
from datetime import datetime
from openpyxl import load_workbook
import os
import mysql.connector
import re


def connect_ftp(host, username, password):
    """Connect to an FTP server and return the connection object."""
    ftp = FTP(host)
    ftp.login(username, password)
    return ftp


def get_most_recent_filename(ftp, prefix):
    """Get the most recent filename from the FTP server matching the given prefix using regular expressions."""
    files = ftp.nlst()
    date_pattern = re.compile(r'\b(\d{2}-\d{2}-\d{4})\.xlsx\b')

    valid_files = []
    for f in files:
        if f.startswith(prefix):
            match = date_pattern.search(f)
            if match:
                try:
                    datetime.strptime(match.group(1), '%d-%m-%Y')
                    valid_files.append(f)
                except ValueError:
                    continue

    if not valid_files:
        return None

    valid_files.sort(key=lambda x: datetime.strptime(date_pattern.search(x).group(1), '%d-%m-%Y'), reverse=True)
    return valid_files[0]


def download_file(ftp, filename, local_path):
    """Download a file from the FTP server to a local path if it does not exist or is different from the base file."""
    if not os.path.exists(local_path) or os.path.basename(local_path) != filename:
        with open(local_path, 'wb') as file:
            ftp.retrbinary(f'RETR {filename}', file.write)
        return True  # Indicates a download occurred
    return False  # Indicates no download was necessary


def connect_db():
    """Connect to a MySQL database and return the connection object."""
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="asesoria_digital"
    )


def insert_data_to_db(data):
    """Insert data into the `asesorias_digitales` table in the `asesoria_digital` database."""
    try:
        db_conn = connect_db()
        cursor = db_conn.cursor()
        insert_query = """
        INSERT INTO asesorias_digitales(Canal, Solicitud, Info_Adicional, Tipo_Documento, Identificacion, Nombres, Apellidos, Asociado, Monto, Correo, Celular, Ocupacion, Otra_Ocupacion, Agencia, Fecha_Solicitud)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.executemany(insert_query, data)
        db_conn.commit()
    except mysql.connector.Error as err:
        print("Error during data insertion:", err)
    finally:
        if db_conn.is_connected():
            cursor.close()
            db_conn.close()


def concatenate_files(base_file_path, new_file_path, insert_into_db=False):
    """Concatenate new data from an Excel file to a base file and optionally insert into the database."""
    base_wb = load_workbook(base_file_path)
    base_ws = base_wb.active

    new_wb = load_workbook(new_file_path)
    new_ws = new_wb.active

    data_to_insert = []
    for row in new_ws.iter_rows(min_row=2, values_only=True):
        base_ws.append(row)
        if insert_into_db:
            data_to_insert.append(row[1:16])  # Exclude the first column ('Ticket')

    base_wb.save(base_file_path)

    if insert_into_db and data_to_insert:
        insert_data_to_db(data_to_insert)


# Configuration
ftp_host = 'ejemplo.com.co'
ftp_username = 'email@ejemplo.com.co'
ftp_password = 'ejemplo1234'
local_download_path = 'C:/Users/miguelcatano.3/PycharmProjects/XXXX EJEMPLO'
base_file_path = os.path.join(local_download_path, 'palabra de referencia aquí')
file_prefix = 'ejemplo del archivo123'

# Connect to the FTP server
ftp_conn = connect_ftp(ftp_host, ftp_username, ftp_password)

# Get the most recent filename and download it if necessary
filename_to_download = get_most_recent_filename(ftp_conn, file_prefix)
if filename_to_download:
    download_path = os.path.join(local_download_path, filename_to_download)

    if not os.path.exists(base_file_path):
        if download_file(ftp_conn, filename_to_download, download_path):
            os.rename(download_path, base_file_path)
            # Insert data into the database
            concatenate_files(base_file_path, base_file_path, insert_into_db=True)
    else:
        if download_file(ftp_conn, filename_to_download, download_path):
            concatenate_files(base_file_path, download_path, insert_into_db=True)
            os.remove(download_path)

# Close the FTP connection
ftp_conn.quit()
